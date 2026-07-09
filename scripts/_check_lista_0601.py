"""Análisis puntual lista vs CSV/Excel 06012026."""
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from cobranzas.domain.services.cartera_merge_service import CarteraMergeService
from cobranzas.domain.services.dias_habiles_service import (
    calcular_cuota_mora,
    fecha_consulta_mora,
)
from cobranzas.domain.services.mora_temprana_service import (
    cuotas_atraso_camorosico,
    dia_pago_desde_credito,
    dias_atraso_camorosico,
    fecha_ultimo_pago_desde_credito,
)
from cobranzas.infrastructure.adapters.cuadro_morosidad_parser import (
    leer_cuadro_morosidad,
)
from cobranzas.infrastructure.adapters.te_detallado_cartera_parser import (
    leer_te_detallado_cartera,
)
from cobranzas.infrastructure.config.settings import Settings
from cobranzas.infrastructure.persistence.database import create_engine_from_settings
from cobranzas.infrastructure.persistence.repositories.feriados_calendario_repository import (
    SqlAlchemyFeriadosCalendarioRepository,
)
from cobranzas.infrastructure.persistence.session import get_session_factory

OPS_NO_EXCEL = [
    "0210002445",
    "0160000176",
    "0210031164",
    "0015336701",
    "0170021704",
    "0018856740",
]
SAMPLE = ["0018645311", "0030220651", "0280029723"]

xlsx = Path("destino/2026/06/asignacion_acumulado_202606.xlsx")
wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
idx_fecha = headers.index("FECHA DEL PROCESO")
idx_op = headers.index("OPERACION")

print("=== Fechas en Excel (muestra elegibles) ===")
for row in ws.iter_rows(min_row=2, values_only=True):
    op = str(row[idx_op] or "").strip().zfill(10)
    if op in [s.zfill(10) for s in SAMPLE]:
        f = row[idx_fecha]
        if hasattr(f, "date"):
            f = f.date()
        print(op, f)
wb.close()

fecha = date(2026, 6, 1)
lote = Path("docsmora/2026/06012026/cartera06012026b")
mor = next(lote.glob("camorosico*.lis"))
car = next(lote.glob("cadetacaco*.lis"))
_, _, morosidad = leer_cuadro_morosidad(mor, fecha_corte_override=fecha)
_, _, cartera = leer_te_detallado_cartera(car, fecha_corte_override=fecha)
mor_por = {c.id_credito.strip(): c for c in morosidad}
car_por = {c.id_credito.strip(): c for c in cartera}
merge = CarteraMergeService()
cfg = Settings()
sf = get_session_factory(create_engine_from_settings(cfg))
feriados = SqlAlchemyFeriadosCalendarioRepository(
    sf, cfg.clave_feriados
).fechas_vigentes()
cons = fecha_consulta_mora(fecha, feriados)

print("\n=== Excluidos detalle (no en CSV ni Excel) ===")
for op in OPS_NO_EXCEL:
    cm = mor_por.get(op)
    if not cm:
        print(op, "NO en CAMOROSICO")
        continue
    cc = car_por.get(op)
    c = merge._merge(cm, cc) if cc else cm
    dp = dia_pago_desde_credito(c)
    up = fecha_ultimo_pago_desde_credito(c, fecha_limite=cons)
    cuota = calcular_cuota_mora(cons, dp, feriados, up)
    print(
        op,
        f"camorosico={dias_atraso_camorosico(c)}",
        f"cuotas_atr={cuotas_atraso_camorosico(c)}",
        f"clasif={cuota.clasificacion}",
        f"dias={cuota.dias}",
        f"mes_cuota={cuota.mes_cuota}",
        f"cuotas_venc={cuota.cuotas_vencidas_impagas}",
        f"up={up}",
    )
