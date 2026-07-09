"""Escritura incremental del Excel acumulado mensual."""

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook, load_workbook

from cobranzas.domain.models.fila_acumulado_mensual import FilaAcumuladoMensual
from cobranzas.domain.ports.acumulado_excel_port import AcumuladoExcelPort

COLUMNAS: tuple[str, ...] = (
    "FECHA DEL PROCESO",
    "DESC OFICINA",
    "NOMBRE",
    "CEDULA",
    "OPERACION",
    "TIPO OPER",
    "FECHA DE CONCESION",
    "VALOR ORI PRESTAMO",
    "SALDO CAP PREST",
    "CALIFICAC",
    "TOTAL PROVISION",
    "SALDO 14,0X",
    "SALDO 14,1X",
    "SALDO 14,2X",
    "TOTAL OP",
    "EST",
    "ESTADO MORA",
    "DIAS ATRASO (CAMOROSICO)",
    "DIAS MORA",
    "TIPO",
    "DIA PAGO",
    "VALOR CUOTA",
    "CUOTA ACT",
    "DIVIDENDOS",
    "OFICIAL ADM",
    "DECISION",
    "SEGMENTACION",
    "SCORE",
    "FUENTE REPAGO",
    "IDENTIFICACION IFI",
    "USUARIO",
    "NOMBRE ASESOR",
    "ID_CREDITO RECBLUE",
)

_IDX_FECHA_PROCESO = COLUMNAS.index("FECHA DEL PROCESO")
_IDX_OPERACION = COLUMNAS.index("OPERACION")


def _formatear_fecha(valor: Optional[date]) -> str:
    if valor is None:
        return ""
    return valor.strftime("%m/%d/%Y")


def _formatear_decimal(valor: Optional[Decimal]) -> object:
    if valor is None:
        return ""
    return float(valor)


def _fila_a_valores(fila: FilaAcumuladoMensual) -> list[object]:
    return [
        _formatear_fecha(fila.fecha_proceso),
        fila.desc_oficina,
        fila.nombre,
        fila.cedula,
        fila.operacion,
        fila.tipo_oper,
        _formatear_fecha(fila.fecha_concesion),
        _formatear_decimal(fila.valor_ori_prestamo),
        _formatear_decimal(fila.saldo_cap_prest),
        fila.calificac,
        _formatear_decimal(fila.total_provision),
        _formatear_decimal(fila.saldo_140x),
        _formatear_decimal(fila.saldo_141x),
        _formatear_decimal(fila.saldo_142x),
        _formatear_decimal(fila.total_op),
        fila.est,
        fila.estado_mora,
        fila.dias_atraso_camorosico if fila.dias_atraso_camorosico is not None else "",
        fila.dias_mora if fila.dias_mora is not None else "",
        fila.tipo,
        fila.dia_pago if fila.dia_pago is not None else "",
        _formatear_decimal(fila.valor_cuota),
        fila.cuota_act if fila.cuota_act is not None else "",
        fila.dividendos if fila.dividendos is not None else "",
        fila.oficial_adm,
        fila.decision,
        fila.segmentacion,
        fila.score,
        fila.fuente_repago,
        fila.identificacion_ifi,
        fila.usuario_asesor,
        fila.nombre_asesor,
        fila.id_credito_recblue,
    ]


def _parsear_fecha_celda(valor) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()[:10]
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _mapa_por_operacion(hoja, columnas: Sequence[str]) -> dict[str, list[object]]:
    """Carga filas existentes indexadas por OPERACION (sin duplicar)."""
    if hoja.max_row < 2:
        return {}
    encabezados = [str(c.value or "").strip() for c in hoja[1]]
    if encabezados != list(columnas):
        return {}

    por_operacion: dict[str, list[object]] = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila or all(v is None or v == "" for v in fila):
            continue
        operacion = str(fila[_IDX_OPERACION] or "").strip()
        if operacion:
            por_operacion[operacion] = list(fila)
    return por_operacion


def _ordenar_filas(filas: Sequence[Sequence[object]]) -> list[list[object]]:
    def clave(fila: Sequence[object]) -> tuple:
        fecha = _parsear_fecha_celda(fila[_IDX_FECHA_PROCESO])
        operacion = str(fila[_IDX_OPERACION] or "").strip()
        return (fecha or date.min, operacion)

    return sorted((list(f) for f in filas), key=clave)


class ExcelAcumuladoWriter(AcumuladoExcelPort):
    def anexar_lote(
        self,
        archivo: Path,
        fecha_corte: date,
        filas: List[FilaAcumuladoMensual],
    ) -> int:
        archivo.parent.mkdir(parents=True, exist_ok=True)

        por_operacion: dict[str, list[object]] = {}
        if archivo.is_file():
            libro = load_workbook(archivo, read_only=True, data_only=True)
            por_operacion = _mapa_por_operacion(libro.active, COLUMNAS)
            libro.close()

        for fila in filas:
            operacion = fila.operacion.strip()
            if not operacion:
                continue
            existente = por_operacion.get(operacion)
            if existente is not None:
                fecha_existente = _parsear_fecha_celda(existente[_IDX_FECHA_PROCESO])
                fecha_nueva = fila.fecha_proceso
                if (
                    fecha_existente is not None
                    and fecha_nueva is not None
                    and fecha_nueva > fecha_existente
                ):
                    # La operación ya estaba registrada en una fecha de proceso
                    # anterior (se conserva del corte previo): no se mueve hacia
                    # adelante ni se duplica.
                    continue
            por_operacion[operacion] = _fila_a_valores(fila)

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Acumulado"
        hoja.append(list(COLUMNAS))
        for fila in _ordenar_filas(por_operacion.values()):
            hoja.append(fila)

        libro.save(archivo)
        return len(filas)
