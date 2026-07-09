import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from cobranzas.domain.models.asesor_registro import AsesorRegistro
from cobranzas.domain.ports.asesor_excel_repository import AsesorExcelRepositoryPort
from cobranzas.infrastructure.persistence.mappers.cobranza_credito_mapper import (
    PREFIJO_CEDULA_ASESOR,
)

ALIAS_COLUMNAS: Dict[str, Tuple[str, ...]] = {
    "cedula": (
        "cedula",
        "cédula",
        "documento",
        "codigo_oficial",
        "codigo",
        "cod_oficial",
        "oficial",
        "id_oficial",
        "usuario",
    ),
    "nombre": (
        "nombre",
        "nombre_asesor",
        "nombre_oficial",
        "asesor",
    ),
    "orden": ("orden", "order", "prioridad"),
    "numero_telefono": (
        "numero_telefono",
        "telefono",
        "teléfono",
        "celular",
        "movil",
        "móvil",
    ),
    "email": ("email", "correo", "mail"),
    "activo": ("activo", "estado", "habilitado"),
}


def _normalizar_encabezado(valor: object) -> str:
    texto = str(valor or "").strip().lower()
    texto = re.sub(r"[^\w]+", "_", texto, flags=re.UNICODE)
    return re.sub(r"_+", "_", texto).strip("_")


def _mapear_encabezados(fila_encabezado: Sequence[object]) -> Dict[str, int]:
    indices: Dict[str, int] = {}
    for idx, celda in enumerate(fila_encabezado):
        clave = _normalizar_encabezado(celda)
        if not clave:
            continue
        for campo, alias in ALIAS_COLUMNAS.items():
            if clave in alias and campo not in indices:
                indices[campo] = idx
    return indices


def _celda(fila: Sequence[object], indice: Optional[int]) -> str:
    if indice is None or indice >= len(fila):
        return ""
    valor = fila[indice]
    if valor is None:
        return ""
    return str(valor).replace("\xa0", " ").strip()


def _parse_orden(valor: object, numero_fila: int) -> Tuple[int, int]:
    try:
        return int(float(str(valor or "").strip())), numero_fila
    except (TypeError, ValueError):
        return numero_fila, numero_fila


def normalizar_cedula_asesor(valor: str) -> str:
    texto = (valor or "").strip().upper()
    if not texto:
        return ""
    if texto.startswith(PREFIJO_CEDULA_ASESOR):
        return texto
    solo_digitos = re.sub(r"\D", "", texto)
    if solo_digitos and solo_digitos == re.sub(r"\D", "", texto):
        return f"{PREFIJO_CEDULA_ASESOR}{solo_digitos.lstrip('0') or solo_digitos}"
    return f"{PREFIJO_CEDULA_ASESOR}{texto}"


def _parse_activo(valor: str) -> bool:
    texto = (valor or "").strip().lower()
    if not texto:
        return True
    if texto in ("0", "no", "false", "inactivo", "n"):
        return False
    if texto in ("1", "si", "sí", "true", "activo", "s", "yes", "y"):
        return True
    return True


class ExcelAsesorReader(AsesorExcelRepositoryPort):
    def leer_asesores(self, archivo_excel: Path) -> List[AsesorRegistro]:
        if not archivo_excel.is_file():
            raise FileNotFoundError(f"No existe Excel de asesores: {archivo_excel}")

        libro = load_workbook(archivo_excel, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:
            raise ValueError(f"Excel vacío: {archivo_excel}")

        indices = _mapear_encabezados(filas[0])
        if "cedula" not in indices:
            raise ValueError(
                "Excel debe tener columna cedula, codigo_oficial o usuario. "
                f"Encabezados detectados: {list(indices.keys())}"
            )

        registros_ordenados: List[Tuple[Tuple[int, int], AsesorRegistro]] = []
        for numero_fila, fila in enumerate(filas[1:], start=2):
            codigo_raw = _celda(fila, indices.get("cedula"))
            nombre = _celda(fila, indices.get("nombre")) or codigo_raw
            cedula = normalizar_cedula_asesor(codigo_raw)
            if not cedula and not nombre:
                continue
            if not cedula:
                raise ValueError(f"Fila {numero_fila}: falta cédula/código de oficial")
            if not nombre:
                raise ValueError(f"Fila {numero_fila}: falta nombre del asesor")

            clave_orden = _parse_orden(
                fila[indices["orden"]] if "orden" in indices else None,
                numero_fila,
            )
            registros_ordenados.append(
                (
                    clave_orden,
                    AsesorRegistro(
                        cedula=cedula,
                        nombre=nombre,
                        numero_telefono=_celda(fila, indices.get("numero_telefono")),
                        email=_celda(fila, indices.get("email")),
                        activo=_parse_activo(_celda(fila, indices.get("activo"))),
                    ),
                )
            )

        if not registros_ordenados:
            raise ValueError(f"Sin filas de datos en {archivo_excel}")

        registros_ordenados.sort(key=lambda par: par[0])
        return [registro for _, registro in registros_ordenados]
