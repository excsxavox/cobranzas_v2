"""Lee export Recblue (Excel .xlsx o CSV) — Consulta Créditos."""

import csv
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from cobranzas.domain.ports.recblue_port import RecbluePort

logger = logging.getLogger("cobranzas.recblue")

CLAVES_OPERACION = (
    "numero_operacion",
    "numero_de_operacion",
    "número_operación",
    "no_operacion",
    "no.operacion",
    "operacion",
)
CLAVES_ID_CREDITO = (
    "id_credito",
    "id_crédito",
    "id_credito_recblue",
    "id",
)
CLAVES_IDENTIFICACION = (
    "identificacion",
    "identificación",
    "identificacion_socio",
    "identificación_socio",
    "cedula",
    "cédula",
)
CLAVES_SOCIO = (
    "socio",
    "codigo_socio",
    "código_socio",
)

COLUMNAS_REQUERIDAS = ("id_credito", "numero_operacion")


def _normalizar_clave(nombre: str) -> str:
    texto = (nombre or "").strip().lower()
    for viejo, nuevo in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        texto = texto.replace(viejo, nuevo)
    return texto.replace(" ", "_")


def _resolver_columna(columnas: Dict[str, str], claves: Tuple[str, ...]) -> Optional[str]:
    return next((columnas[k] for k in claves if k in columnas), None)


class RecblueArchivoAdapter(RecbluePort):
    """Carga ID Crédito por número de operación desde export Recblue."""

    def __init__(self, archivo: Optional[Path]) -> None:
        self._archivo = archivo
        self._ultimos_errores: List[str] = []

    @property
    def errores_validacion(self) -> List[str]:
        return list(self._ultimos_errores)

    def id_credito_por_operacion(self) -> Dict[str, str]:
        self._ultimos_errores = []
        if self._archivo is None or not self._archivo.is_file():
            return {}

        sufijo = self._archivo.suffix.lower()
        if sufijo in (".xlsx", ".xlsm"):
            filas = self._leer_excel(self._archivo)
        elif sufijo == ".csv":
            filas = self._leer_csv(self._archivo)
        else:
            self._ultimos_errores.append(
                f"Formato no soportado para Recblue: {self._archivo.suffix} "
                "(use .xlsx o .csv)"
            )
            return {}

        return self._construir_mapa(filas)

    def _leer_excel(self, ruta: Path) -> List[Dict[str, str]]:
        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas_iter = hoja.iter_rows(values_only=True)
        try:
            encabezados_raw = next(filas_iter)
        except StopIteration:
            self._ultimos_errores.append(f"Excel Recblue vacío: {ruta}")
            libro.close()
            return []

        encabezados = [str(c or "").strip() for c in encabezados_raw]
        columnas = {_normalizar_clave(h): h for h in encabezados if h}

        registros: List[Dict[str, str]] = []
        for row in filas_iter:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            item = {}
            for idx, header in enumerate(encabezados):
                if not header:
                    continue
                valor = row[idx] if idx < len(row) else None
                item[header] = "" if valor is None else str(valor).strip()
            registros.append(item)

        libro.close()
        self._validar_columnas(columnas, encabezados)
        return registros

    def _leer_csv(self, ruta: Path) -> List[Dict[str, str]]:
        with ruta.open(encoding="utf-8-sig", newline="") as fh:
            lector = csv.DictReader(fh)
            if not lector.fieldnames:
                self._ultimos_errores.append(f"CSV Recblue sin encabezados: {ruta}")
                return []
            columnas = {_normalizar_clave(c): c for c in lector.fieldnames}
            self._validar_columnas(columnas, list(lector.fieldnames))
            return [dict(fila) for fila in lector]

    def _validar_columnas(
        self, columnas: Dict[str, str], originales: List[str]
    ) -> None:
        normalizadas = set(columnas.keys())
        if not _resolver_columna(columnas, CLAVES_ID_CREDITO):
            self._ultimos_errores.append(
                "Recblue: falta columna 'ID Crédito' (id_credito). "
                f"Columnas: {originales}"
            )
        if not _resolver_columna(columnas, CLAVES_OPERACION):
            self._ultimos_errores.append(
                "Recblue: falta columna 'Número Operación' (numero_operacion). "
                f"Columnas: {originales}"
            )

    def _construir_mapa(self, filas: List[Dict[str, str]]) -> Dict[str, str]:
        if self._ultimos_errores:
            logger.error("Recblue inválido: %s", "; ".join(self._ultimos_errores))
            return {}

        if not filas:
            return {}

        columnas = {_normalizar_clave(k): k for k in filas[0].keys()}
        col_op = _resolver_columna(columnas, CLAVES_OPERACION)
        col_id = _resolver_columna(columnas, CLAVES_ID_CREDITO)
        if not col_op or not col_id:
            return {}

        mapa: Dict[str, str] = {}
        duplicados = 0
        for fila in filas:
            operacion = (fila.get(col_op) or "").strip()
            id_credito = (fila.get(col_id) or "").strip()
            if not operacion or not id_credito:
                continue
            if operacion in mapa and mapa[operacion] != id_credito:
                duplicados += 1
            mapa[operacion] = id_credito

        if duplicados:
            logger.warning(
                "Recblue: %s operaciones con ID Crédito duplicado/conflictivo",
                duplicados,
            )

        logger.info(
            "Recblue cargado: %s operaciones | archivo=%s",
            len(mapa),
            self._archivo,
        )
        return mapa

    def operaciones_por_id_credito(self, id_credito: str) -> List[Dict[str, str]]:
        """Operaciones core vinculadas a un ID Crédito Recblue (puede haber más de una)."""
        id_buscado = (id_credito or "").strip()
        if not id_buscado:
            return []
        mapa = self.id_credito_por_operacion()
        return [
            {"numero_operacion": op, "id_credito_recblue": id_rb}
            for op, id_rb in mapa.items()
            if id_rb.replace(".0", "") == id_buscado.replace(".0", "")
        ]

    def registro_por_operacion(self, numero_operacion: str) -> Optional[Dict[str, str]]:
        """Fila Recblue normalizada (id, operación, identificación, socio)."""
        if self._archivo is None or not self._archivo.is_file():
            return None
        mapa = self.id_credito_por_operacion()
        id_credito = mapa.get(numero_operacion)
        if not id_credito:
            return None
        return {"id_credito_recblue": id_credito, "numero_operacion": numero_operacion}
