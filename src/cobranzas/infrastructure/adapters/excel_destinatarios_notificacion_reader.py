import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from cobranzas.domain.models.destinatario_notificacion import DestinatarioNotificacion
from cobranzas.domain.ports.destinatarios_notificacion_port import (
    DestinatariosNotificacionPort,
)

ALIAS_COLUMNAS: Dict[str, Tuple[str, ...]] = {
    "nombre": ("nombre", "nombre_contacto", "contacto", "responsable"),
    "email": ("email", "correo", "mail", "e_mail"),
    "activo": ("activo", "estado", "habilitado", "recibe_alertas"),
}


def _normalizar_encabezado(valor: object) -> str:
    texto = str(valor or "").strip().lower()
    texto = re.sub(r"[^\w]+", "_", texto, flags=re.UNICODE)
    return re.sub(r"_+", "_", texto).strip("_")


def _mapear_encabezados(fila_encabezado: Sequence[object]) -> Dict[str, int]:
    indices: Dict[str, int] = {}
    for idx, celda in enumerate(fila_encabezado):
        clave = _normalizar_encabezado(celda)
        if not clave:
            continue
        for campo, alias in ALIAS_COLUMNAS.items():
            if clave in alias and campo not in indices:
                indices[campo] = idx
    return indices


def _celda(fila: Sequence[object], indice: Optional[int]) -> str:
    if indice is None or indice >= len(fila):
        return ""
    valor = fila[indice]
    if valor is None:
        return ""
    return str(valor).strip()


def _parse_activo(valor: str) -> bool:
    texto = (valor or "").strip().lower()
    if not texto:
        return True
    if texto in ("0", "no", "false", "inactivo", "n"):
        return False
    return True


class ExcelDestinatariosNotificacionReader(DestinatariosNotificacionPort):

    def leer_destinatarios(self, archivo_excel: Path) -> List[DestinatarioNotificacion]:
        if not archivo_excel.is_file():
            raise FileNotFoundError(
                f"No existe Excel de notificaciones: {archivo_excel}"
            )

        libro = load_workbook(archivo_excel, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:
            
            raise ValueError(f"Excel vacío: {archivo_excel}")

        indices = _mapear_encabezados(filas[0])
        if "email" not in indices:
            raise ValueError(
                "Excel debe tener columna email (o correo). "
                f"Encabezados detectados: {list(indices.keys())}"
            )

        registros: List[DestinatarioNotificacion] = []
        for fila in filas[1:]:
            email = _celda(fila, indices.get("email"))
            nombre = _celda(fila, indices.get("nombre")) or email

            if not email:
                continue
            registros.append(
                DestinatarioNotificacion(
                    nombre=nombre,
                    email=email,
                    activo=_parse_activo(_celda(fila, indices.get("activo"))),
                )
            )

        if not registros:
            raise ValueError(f"Sin filas de datos en {archivo_excel}")

        return registros
