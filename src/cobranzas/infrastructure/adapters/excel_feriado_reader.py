import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from cobranzas.domain.models.feriado_rango import FeriadoRango
from cobranzas.domain.ports.feriado_excel_repository import FeriadoExcelRepositoryPort
from cobranzas.domain.services.feriado_fechas import parsear_fecha_excel

log = logging.getLogger("cobranzas.sync.feriados.excel")


def _normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


class ExcelFeriadoReader(FeriadoExcelRepositoryPort):
    def buscar_archivo(self, directorio: Path, patron: str) -> Optional[Path]:
        if not directorio.exists():
            log.warning("No existe el directorio de feriados: %s", directorio)
            return None

        archivos = [
            p
            for p in directorio.glob(patron)
            if p.is_file() and not p.name.startswith("~$")
        ]
        if not archivos:
            return None

        archivos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        if len(archivos) > 1:
            log.info(
                "Se encontraron %s archivos de feriados. Se usará el más reciente: %s",
                len(archivos),
                archivos[0],
            )
        return archivos[0]

    def leer_feriados(self, archivo_excel: Path) -> List[FeriadoRango]:
        if not archivo_excel.is_file():
            raise FileNotFoundError(f"No existe Excel de feriados: {archivo_excel}")

        log.info("Leyendo feriados desde: %s", archivo_excel)
        libro = load_workbook(archivo_excel, read_only=True, data_only=True)
        hoja = libro.active

        registros: List[FeriadoRango] = []
        numero_fila = 0

        for row in hoja.iter_rows(values_only=True):
            numero_fila += 1
            valores = [v for v in row if v is not None and str(v).strip() != ""]
            if not valores:
                continue

            fechas = []
            textos = []
            for valor in valores:
                fecha = parsear_fecha_excel(valor)
                if fecha is not None:
                    fechas.append(fecha)
                else:
                    texto = _normalizar_texto(valor)
                    if texto:
                        textos.append(texto)

            if not textos:
                log.warning(
                    "Fila %s omitida: sin descripción. valores=%s",
                    numero_fila,
                    valores,
                )
                continue

            descripcion = textos[0]

            if len(fechas) == 2:
                fecha_inicio = min(fechas)
                fecha_fin = max(fechas)
            elif len(fechas) == 1:
                fecha_inicio = fechas[0]
                fecha_fin = fechas[0]
                log.info(
                    "Fila %s: feriado de un día | descripcion='%s' | fecha=%s",
                    numero_fila,
                    descripcion,
                    fecha_inicio,
                )
            else:
                log.warning(
                    "Fila %s omitida: sin fecha válida. valores=%s",
                    numero_fila,
                    valores,
                )
                continue

            registros.append(
                FeriadoRango(
                    descripcion=descripcion,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                )
            )

        libro.close()

        if not registros:
            raise RuntimeError(
                f"No se obtuvo ningún feriado válido desde {archivo_excel}"
            )

        return registros
