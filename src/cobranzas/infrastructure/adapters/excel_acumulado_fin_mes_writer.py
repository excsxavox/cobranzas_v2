"""Escritura incremental del Excel acumulado fin de mes (deudores + deuda)."""

from dataclasses import fields
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook, load_workbook

from cobranzas.domain.models.fila_acumulado_fin_mes import FilaAcumuladoFinMes
from cobranzas.domain.ports.acumulado_fin_mes_excel_port import AcumuladoFinMesExcelPort

ENCABEZADOS: tuple[str, ...] = (
    "FECHA DEL PROCESO",
    "DEUDORES_NOMBRE",
    "DEUDORES_DOCUMENTO",
    "DEUDORES_SOCIO",
    "NUMERO_OPERACION",
    "FECHA_CORTE",
    "ARCHIVO_ORIGEN",
    "OFICINA",
    "DESC_OFICINA",
    "SOCIO",
    "NOMBRE",
    "CEDULA",
    "SECTOR",
    "TIPO_OPERACION",
    "TIPO_DESTINO",
    "FECHA_CONCESION",
    "FECHA_VENCIMIENTO",
    "FECHA_ULTIMO_PAGO",
    "VALOR_ORIGINAL_PRESTAMO",
    "SALDO_CAPITAL_PRESTAMO",
    "CALIFICACION",
    "TOTAL_PROVISION",
    "SALDO_140X",
    "SALDO_141X",
    "SALDO_142X",
    "INTERES_NORMAL",
    "INTERES_DEVENGADO",
    "INTERES_VENCIDO",
    "INTERES_RESOLUCION",
    "INTERES_CASTIGADO",
    "INTERES_MORA",
    "OTROS_RUBROS_DEUDA",
    "TOTAL_OPERACION",
    "ESTADO",
    "OFICIAL",
    "DIAS_MORA",
    "DIAS_ATRASO_CAMOROSICO",
    "FECHA_INGRESO",
    "TIPO",
    "DIA_PAGO",
    "VALOR_CUOTA",
    "CUOTA_ACTUAL",
    "DIVIDENDOS",
    "COD_OFICIAL_ASIGNADO",
    "OFICIAL_ASIGNADO",
    "COD_OFICIAL_ADM",
    "OFICIAL_ADM",
    "OPERACION_HOMOLOGADA",
    "DECISION",
    "SEGMENTACION",
    "SCORE",
    "FUENTE_REPAGO",
    "IDENTIFICACION_IFI",
    "ACTIVIDAD_ECONOMICA",
    "FECHA_ARCHIVO",
    "TIPO_MES",
    "TIPO_FIDEICOMISO",
    "PROCESO_COD",
    "ESTADO_MORA",
    "ID_CREDITO_RECBLUE",
)

_IDX_FECHA_PROCESO = ENCABEZADOS.index("FECHA DEL PROCESO")
_IDX_NUMERO_OPERACION = ENCABEZADOS.index("NUMERO_OPERACION")

_CAMPOS_DATACLASS = tuple(f.name for f in fields(FilaAcumuladoFinMes))


def _formatear_fecha(valor: Optional[date]) -> str:
    if valor is None:
        return ""
    return valor.strftime("%m/%d/%Y")


def _formatear_decimal(valor: Optional[Decimal]) -> object:
    if valor is None:
        return ""
    return float(valor)


def _formatear_valor(nombre_campo: str, valor) -> object:
    if valor is None:
        return ""
    if isinstance(valor, Decimal):
        return _formatear_decimal(valor)
    if isinstance(valor, (date, datetime)):
        if isinstance(valor, datetime):
            return _formatear_fecha(valor.date())
        return _formatear_fecha(valor)
    return valor


def _fila_a_valores(fila: FilaAcumuladoFinMes) -> list[object]:
    return [
        _formatear_valor(nombre, getattr(fila, nombre)) for nombre in _CAMPOS_DATACLASS
    ]


def _parsear_fecha_celda(valor) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()[:10]
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _mapa_por_operacion(hoja, columnas: Sequence[str]) -> dict[str, list[object]]:
    if hoja.max_row < 2:
        return {}
    encabezados = [str(c.value or "").strip() for c in hoja[1]]
    if encabezados != list(columnas):
        return {}

    por_operacion: dict[str, list[object]] = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila or all(v is None or v == "" for v in fila):
            continue
        operacion = str(fila[_IDX_NUMERO_OPERACION] or "").strip()
        if operacion:
            por_operacion[operacion] = list(fila)
    return por_operacion


def _ordenar_filas(filas: Sequence[Sequence[object]]) -> list[list[object]]:
    def clave(fila: Sequence[object]) -> tuple:
        fecha = _parsear_fecha_celda(fila[_IDX_FECHA_PROCESO])
        operacion = str(fila[_IDX_NUMERO_OPERACION] or "").strip()
        return (fecha or date.min, operacion)

    return sorted((list(f) for f in filas), key=clave)


class ExcelAcumuladoFinMesWriter(AcumuladoFinMesExcelPort):
    def anexar_lote(
        self,
        archivo: Path,
        fecha_corte: date,
        filas: List[FilaAcumuladoFinMes],
    ) -> int:
        archivo.parent.mkdir(parents=True, exist_ok=True)

        por_operacion: dict[str, list[object]] = {}
        if archivo.is_file():
            libro = load_workbook(archivo, read_only=True, data_only=True)
            por_operacion = _mapa_por_operacion(libro.active, ENCABEZADOS)
            libro.close()

        for fila in filas:
            operacion = fila.numero_operacion.strip()
            if operacion:
                por_operacion[operacion] = _fila_a_valores(fila)

        libro = Workbook()
        hoja = libro.active
        hoja.title = "AcumuladoFinMes"
        hoja.append(list(ENCABEZADOS))
        for fila in _ordenar_filas(por_operacion.values()):
            hoja.append(fila)

        libro.save(archivo)
        return len(filas)
