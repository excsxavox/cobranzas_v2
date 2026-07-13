"""
Genera reportes Excel de gestión preventiva (HU GRC-03 líneas 275-308).

`escribir_reporte_mensual` genera el reporte consolidado por corte o mensual
con TODOS los cortes y gestiones del período (HU líneas 275-284).
"""

import logging
from pathlib import Path
from typing import List, TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from preventiva.infrastructure.persistence.models.reporte_preventiva import ReportePreventiva

log = logging.getLogger("preventiva.adapters.reporte_excel")

CABECERAS = [
    "Fecha Proceso", "Nombre", "Cédula", "Numero Operación",
    "Días Mora", "Día Pago", "Teléfono Celular",
    "Saldo Pendiente Cuota", "Saldo en Cuenta", "Cobertura", "Gestión N°",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _escribir_cabeceras(ws):
    for col, cab in enumerate(CABECERAS, 1):
        cell = ws.cell(row=1, column=col, value=cab)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def escribir_reporte_mensual(
    filas_bd: "List[ReportePreventiva]",
    directorio: Path,
    anio: int,
    mes: int,
) -> Path:
    """
    Reporte mensual consolidado (HU líneas 275-284).
    Consolida TODOS los cortes y gestiones del mes en un único Excel.
    Nombre: REPORTE_PREVENTIVA_MMAAAA.xlsx  (ej. REPORTE_PREVENTIVA_052026.xlsx).
    Una hoja por corte del mes, más una hoja "Resumen" con todo consolidado.
    """
    nombre = f"REPORTE_PREVENTIVA_{mes:02d}{anio}.xlsx"
    ruta = directorio / nombre
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Hoja de resumen general (primera hoja)
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Mensual"
    _escribir_cabeceras(ws_resumen)

    # Agrupa por corte para crear una hoja por corte
    cortes_vistos: dict = {}
    fila_resumen = 2

    for r in filas_bd:
        fecha_str = r.fecha_proceso.strftime("%d/%m/%Y") if r.fecha_proceso else ""

        # Fila en hoja de resumen
        ws_resumen.cell(row=fila_resumen, column=1,  value=fecha_str)
        ws_resumen.cell(row=fila_resumen, column=2,  value=r.nombre)
        ws_resumen.cell(row=fila_resumen, column=3,  value=r.cedula)
        ws_resumen.cell(row=fila_resumen, column=4,  value=r.numero_operacion)
        ws_resumen.cell(row=fila_resumen, column=5,  value=r.dias_mora)
        ws_resumen.cell(row=fila_resumen, column=6,  value=r.dia_pago)
        ws_resumen.cell(row=fila_resumen, column=7,  value=r.telefono)
        ws_resumen.cell(row=fila_resumen, column=8,  value=float(r.saldo_pendiente or 0))
        ws_resumen.cell(row=fila_resumen, column=9,  value=float(r.saldo_cuenta or 0))
        ws_resumen.cell(row=fila_resumen, column=10, value="")
        ws_resumen.cell(row=fila_resumen, column=11, value=r.numero_gestion)
        fila_resumen += 1

        # Hoja por corte
        corte = r.dia_corte or 0
        if corte not in cortes_vistos:
            ws_corte = wb.create_sheet(title=f"Corte {corte:02d}")
            _escribir_cabeceras(ws_corte)
            cortes_vistos[corte] = (ws_corte, 2)

        ws_c, fila_c = cortes_vistos[corte]
        ws_c.cell(row=fila_c, column=1,  value=fecha_str)
        ws_c.cell(row=fila_c, column=2,  value=r.nombre)
        ws_c.cell(row=fila_c, column=3,  value=r.cedula)
        ws_c.cell(row=fila_c, column=4,  value=r.numero_operacion)
        ws_c.cell(row=fila_c, column=5,  value=r.dias_mora)
        ws_c.cell(row=fila_c, column=6,  value=r.dia_pago)
        ws_c.cell(row=fila_c, column=7,  value=r.telefono)
        ws_c.cell(row=fila_c, column=8,  value=float(r.saldo_pendiente or 0))
        ws_c.cell(row=fila_c, column=9,  value=float(r.saldo_cuenta or 0))
        ws_c.cell(row=fila_c, column=10, value="")
        ws_c.cell(row=fila_c, column=11, value=r.numero_gestion)
        cortes_vistos[corte] = (ws_c, fila_c + 1)

    wb.save(ruta)
    log.info(
        "Reporte mensual: %s (%d filas, %d cortes)",
        ruta.name, len(filas_bd), len(cortes_vistos),
    )
    return ruta
